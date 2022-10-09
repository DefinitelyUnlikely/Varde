# Om vi använder excel, hur gör vi det på bästa sätt? Vi hämtar ut regionerna och gör regionen till en map (dictionary). 
# I regions map:en kommer vi att mappa region till butik. butik till värde och gross. Vad vi kan fundera på är om
# vi också vill ha värdet och grossen direkt i regionen, istället för att det skall räknas ut efteråt. i.e. vi lägger in det direkt istället.

# Sedan skall vi nog fundera på att ha varje dag som sin egen dictionary isåfall. Så att vi på ett enkelt sätt kan jämföra olika tidsperioder?
# Så först vill ajg lösa strukturen för att få indata. Sedan får vi lösa ett bra sätt att skapa en databas. Sedan får vi utnyttja databasen för
# att få fram data för olika perioder.

from openpyxl import load_workbook
from collections import defaultdict

def main():
    pass

    def read_file():
        workbook = load_workbook("C:\Code\Projects\Varde\9Sep.xlsx")
        worksheet = workbook[workbook.sheetnames[1]]
        region_column = worksheet["D"]
        chain_column = worksheet["E"]
        store_column = worksheet["F"]
        value_column = worksheet["G"]
        
        return zip(region_column[1:], chain_column[1:], store_column[1:], value_column[1:])


    value_dict = defaultdict()
    
    zipped = read_file()
    # for i, j, k, l in zipped:
    #     print(i.value.encode("utf-8"), j.value.encode("utf-8"), k.value.encode("utf-8"), l.value)
        
    
    for region, chain, store, value in zipped:
        print(region.value.encode("utf-8"))
        print(chain.value.encode("utf-8"))
        print(store.value.encode("utf-8"))
        print(value.value)






if __name__ == '__main__':
    main() 